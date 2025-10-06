import openpyxl
import csv

book = openpyxl.load_workbook("C:\\Users\\harsh\\OneDrive\\Documents\\tpythonexcel003.xlsx")

sheet= book.active
cell = sheet.cell(row=3, column=2)
print(cell.value)

sheet.cell(row=4, column=4).value = "adi123@gmail.com"

print(sheet.cell(row=4, column=4).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)

# More practice on lecture 63 - udemy for excel learn again