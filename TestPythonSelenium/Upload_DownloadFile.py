import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

#Revisit this section to understand -- 66 number session

def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)


fruit_name = "Apple"
#file_path = "C:\\Users\\harsh\\Downloads\\download(1).xlsx"
file_path = "C:\\Users\\harsh\\OneDrive\\Documents\\download_12.xlsx"
new_val = "532"
driver = webdriver.Firefox()
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.implicitly_wait(4)

driver.find_element(By.CSS_SELECTOR, "#downloadButton").click()

#edit excel code
update_excel_data(file_path, fruit_name, "price", new_val)

#upload File
file_up= driver.find_element(By.XPATH, "//input[@type='file']")
file_up.send_keys(file_path)

# if a toast msg comes -- then locate and keep ss to print msg in tosat or locating element.
wait = WebDriverWait(driver, 5)
toast_msg = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_msg))

print(driver.find_element(*toast_msg).text)

price_c = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id = 'cell-"+price_c+"-undefined']").text

#actual_price = driver.find_element(
   # By.XPATH,
  #  f'//div[text()="{fruit_name}"]/parent::div/parent::div/div[@id="{price_c}"]'
#)
assert actual_price == new_val


